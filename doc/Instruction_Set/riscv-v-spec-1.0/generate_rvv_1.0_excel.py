#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RISC-V Vector Extension 1.0 Excel Generator with Diff Info
Generate Excel file from updated JSON with diff information
"""

import json
import os
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: openpyxl not available.")

V10_PATH = r"d:\code\openc910\doc\Instruction_Set\riscv-v-spec-1.0\rvv_instructions.json"
EXCEL_PATH = r"d:\code\openc910\doc\Instruction_Set\riscv-v-spec-1.0\rvv_instructions.xlsx"

def create_excel():
    with open(V10_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    instructions = data['instructions']
    removed = data.get('removed_instructions', [])
    diff_summary = data.get('diff_summary', {})
    
    wb = Workbook()
    
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    removed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    modified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws_summary = wb.active
    ws_summary.title = "差异总览"
    
    ws_summary['A1'] = "RISC-V Vector Extension 1.0 vs 0.7.1 差异总览"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A3'] = "差异统计"
    ws_summary['A3'].font = Font(bold=True, size=12)
    
    ws_summary['A4'] = "差异类型"
    ws_summary['B4'] = "数量"
    ws_summary['C4'] = "占比"
    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}4'].font = header_font_white
        ws_summary[f'{col}4'].fill = header_fill
        ws_summary[f'{col}4'].border = thin_border
    
    total = len(instructions)
    stats = defaultdict(int)
    for inst in instructions:
        stats[inst.get('diff_info', {}).get('type', '未知')] += 1
    
    row = 5
    for diff_type, count in [('新增', stats['新增']), ('修改', stats['修改']), ('无变化', stats['无变化'])]:
        ws_summary[f'A{row}'] = diff_type
        ws_summary[f'B{row}'] = count
        ws_summary[f'C{row}'] = f"{count/total*100:.1f}%"
        for col in ['A', 'B', 'C']:
            ws_summary[f'{col}{row}'].border = thin_border
        row += 1
    
    ws_summary[f'A{row}'] = "删除(0.7.1存在)"
    ws_summary[f'B{row}'] = len(removed)
    ws_summary[f'C{row}'] = "-"
    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}{row}'].border = thin_border
    row += 1
    
    ws_summary[f'A{row}'] = "总计(1.0)"
    ws_summary[f'B{row}'] = total
    ws_summary[f'C{row}'] = "100%"
    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}{row}'].font = Font(bold=True)
        ws_summary[f'{col}{row}'].border = thin_border
    
    row += 2
    ws_summary[f'A{row}'] = "主要变化"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    for change in diff_summary.get('major_changes', []):
        ws_summary[f'A{row}'] = f"• {change}"
        row += 1
    
    ws_summary.column_dimensions['A'].width = 50
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 10
    
    ws_all = wb.create_sheet("指令总表")
    
    headers = ["指令名", "分类", "funct6", "funct3", "指令类型", "操作数", "中文描述", "差异类型", "0.7.1名称", "差异详情"]
    for col, header in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row_idx, inst in enumerate(instructions, 2):
        diff_info = inst.get('diff_info', {})
        diff_type = diff_info.get('type', '')
        
        ws_all.cell(row=row_idx, column=1, value=inst['name']).border = thin_border
        ws_all.cell(row=row_idx, column=2, value=inst.get('category', '')).border = thin_border
        ws_all.cell(row=row_idx, column=3, value=inst.get('funct6', '')).border = thin_border
        ws_all.cell(row=row_idx, column=4, value=inst.get('funct3', '')).border = thin_border
        ws_all.cell(row=row_idx, column=5, value=inst.get('instruction_type', '')).border = thin_border
        ws_all.cell(row=row_idx, column=6, value=inst.get('operands', '')).border = thin_border
        ws_all.cell(row=row_idx, column=7, value=inst.get('description_cn', '')).border = thin_border
        
        diff_cell = ws_all.cell(row=row_idx, column=8, value=diff_type)
        diff_cell.border = thin_border
        if diff_type == "新增":
            diff_cell.fill = added_fill
        elif diff_type == "修改":
            diff_cell.fill = modified_fill
        elif diff_type == "删除":
            diff_cell.fill = removed_fill
        
        ws_all.cell(row=row_idx, column=9, value=diff_info.get('v071_name', '')).border = thin_border
        ws_all.cell(row=row_idx, column=10, value=diff_info.get('details', '')).border = thin_border
    
    ws_all.column_dimensions['A'].width = 15
    ws_all.column_dimensions['B'].width = 25
    ws_all.column_dimensions['C'].width = 10
    ws_all.column_dimensions['D'].width = 8
    ws_all.column_dimensions['E'].width = 25
    ws_all.column_dimensions['F'].width = 20
    ws_all.column_dimensions['G'].width = 30
    ws_all.column_dimensions['H'].width = 10
    ws_all.column_dimensions['I'].width = 15
    ws_all.column_dimensions['J'].width = 50
    
    ws_added = wb.create_sheet("新增指令")
    added_insts = [i for i in instructions if i.get('diff_info', {}).get('type') == '新增']
    
    for col, header in enumerate(headers[:7], 1):
        cell = ws_added.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, inst in enumerate(added_insts, 2):
        for col, key in enumerate(['name', 'category', 'funct6', 'funct3', 'instruction_type', 'operands', 'description_cn'], 1):
            ws_added.cell(row=row_idx, column=col, value=inst.get(key, '')).border = thin_border
    
    for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'], [15, 25, 10, 8, 25, 20, 30]):
        ws_added.column_dimensions[col].width = width
    
    ws_modified = wb.create_sheet("修改指令")
    modified_insts = [i for i in instructions if i.get('diff_info', {}).get('type') == '修改']
    
    for col, header in enumerate(headers, 1):
        cell = ws_modified.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, inst in enumerate(modified_insts, 2):
        diff_info = inst.get('diff_info', {})
        for col, key in enumerate(['name', 'category', 'funct6', 'funct3', 'instruction_type', 'operands', 'description_cn'], 1):
            ws_modified.cell(row=row_idx, column=col, value=inst.get(key, '')).border = thin_border
        ws_modified.cell(row=row_idx, column=8, value=diff_info.get('type', '')).border = thin_border
        ws_modified.cell(row=row_idx, column=9, value=diff_info.get('v071_name', '')).border = thin_border
        ws_modified.cell(row=row_idx, column=10, value=diff_info.get('details', '')).border = thin_border
    
    for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'], [15, 25, 10, 8, 25, 20, 30, 10, 15, 50]):
        ws_modified.column_dimensions[col].width = width
    
    ws_removed = wb.create_sheet("删除指令")
    
    removed_headers = ["指令名", "分类", "funct6", "funct3", "指令类型", "操作数", "中文描述", "1.0对应名称", "说明"]
    for col, header in enumerate(removed_headers, 1):
        cell = ws_removed.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, inst in enumerate(removed, 2):
        diff_info = inst.get('diff_info', {})
        for col, key in enumerate(['name', 'category', 'funct6', 'funct3', 'instruction_type', 'operands', 'description_cn'], 1):
            ws_removed.cell(row=row_idx, column=col, value=inst.get(key, '')).border = thin_border
        ws_removed.cell(row=row_idx, column=8, value=diff_info.get('v10_name', '')).border = thin_border
        ws_removed.cell(row=row_idx, column=9, value=diff_info.get('details', '')).border = thin_border
    
    for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'], [15, 25, 10, 8, 25, 20, 30, 15, 30]):
        ws_removed.column_dimensions[col].width = width
    
    wb.save(EXCEL_PATH)
    print(f"Excel file generated: {EXCEL_PATH}")
    print(f"  新增指令: {len(added_insts)}")
    print(f"  修改指令: {len(modified_insts)}")
    print(f"  删除指令: {len(removed)}")

if __name__ == "__main__":
    create_excel()
