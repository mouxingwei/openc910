#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RISC-V Vector Extension (RVV 0.7.1) Instruction Set Excel Generator
Generate Excel file from JSON data
"""

import json
import os
import sys

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: pandas and openpyxl not available. Using basic Excel generation.")

BASE_PATH = r"d:\code\openc910\doc\Instruction_Set\riscv-v-spec-0.7.1"
JSON_FILE = os.path.join(BASE_PATH, "rvv_instructions_0.7.1.json")
EXCEL_FILE = os.path.join(BASE_PATH, "rvv_instructions_0.7.1.xlsx")

def create_excel_with_pandas():
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    instructions = data['instructions']
    categories = data['categories']
    
    wb = Workbook()
    
    ws_summary = wb.active
    ws_summary.title = "分类总览"
    
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws_summary['A1'] = "RISC-V Vector Extension 0.7.1 指令集分类总览"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A3'] = "分类名称"
    ws_summary['B3'] = "英文名称"
    ws_summary['C3'] = "指令数量"
    ws_summary['D3'] = "占比"
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws_summary[f'{col}3']
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    total = data['metadata']['total_instructions']
    row = 4
    for cat_key, cat_info in sorted(categories.items(), key=lambda x: -x[1]['count']):
        ws_summary[f'A{row}'] = cat_info['name']
        ws_summary[f'B{row}'] = cat_key
        ws_summary[f'C{row}'] = cat_info['count']
        ws_summary[f'D{row}'] = f"{cat_info['count']/total*100:.1f}%"
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].border = thin_border
        row += 1
    
    ws_summary[f'A{row}'] = "总计"
    ws_summary[f'C{row}'] = total
    ws_summary[f'D{row}'] = "100%"
    for col in ['A', 'B', 'C', 'D']:
        ws_summary[f'{col}{row}'].font = Font(bold=True)
        ws_summary[f'{col}{row}'].border = thin_border
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 10
    
    ws_all = wb.create_sheet("指令总表")
    
    headers = ["指令名", "分类", "funct6", "funct3", "指令类型", "操作数", "中文描述", "英文描述", "掩码支持", "格式类型"]
    for col, header in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row_idx, inst in enumerate(instructions, 2):
        ws_all.cell(row=row_idx, column=1, value=inst['name']).border = thin_border
        ws_all.cell(row=row_idx, column=2, value=inst['category']).border = thin_border
        ws_all.cell(row=row_idx, column=3, value=inst['funct6']).border = thin_border
        ws_all.cell(row=row_idx, column=4, value=inst['funct3']).border = thin_border
        ws_all.cell(row=row_idx, column=5, value=inst['instruction_type']).border = thin_border
        ws_all.cell(row=row_idx, column=6, value=inst['operands']).border = thin_border
        ws_all.cell(row=row_idx, column=7, value=inst['description_cn']).border = thin_border
        ws_all.cell(row=row_idx, column=8, value=inst['description_en']).border = thin_border
        ws_all.cell(row=row_idx, column=9, value=inst['vm_support']).border = thin_border
        ws_all.cell(row=row_idx, column=10, value=inst['format']).border = thin_border
    
    ws_all.column_dimensions['A'].width = 15
    ws_all.column_dimensions['B'].width = 25
    ws_all.column_dimensions['C'].width = 20
    ws_all.column_dimensions['D'].width = 8
    ws_all.column_dimensions['E'].width = 25
    ws_all.column_dimensions['F'].width = 20
    ws_all.column_dimensions['G'].width = 35
    ws_all.column_dimensions['H'].width = 30
    ws_all.column_dimensions['I'].width = 10
    ws_all.column_dimensions['J'].width = 15
    
    ws_encoding = wb.create_sheet("编码参考")
    
    ws_encoding['A1'] = "RISC-V Vector Extension 0.7.1 指令编码参考"
    ws_encoding['A1'].font = Font(bold=True, size=14)
    ws_encoding.merge_cells('A1:F1')
    
    ws_encoding['A3'] = "funct3 编码"
    ws_encoding['A3'].font = Font(bold=True, size=12)
    
    enc_headers = ["funct3", "指令类型", "说明"]
    for col, header in enumerate(enc_headers, 1):
        cell = ws_encoding.cell(row=4, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    
    op_types = data['instruction_types']
    for row_idx, (funct3, desc) in enumerate(op_types.items(), 5):
        ws_encoding.cell(row=row_idx, column=1, value=funct3).border = thin_border
        ws_encoding.cell(row=row_idx, column=2, value=desc.split('(')[0].strip()).border = thin_border
        ws_encoding.cell(row=row_idx, column=3, value=desc).border = thin_border
    
    ws_encoding.column_dimensions['A'].width = 10
    ws_encoding.column_dimensions['B'].width = 15
    ws_encoding.column_dimensions['C'].width = 30
    
    for cat_key, cat_info in categories.items():
        safe_name = cat_key.replace('-', '_')[:31]
        ws_cat = wb.create_sheet(safe_name)
        
        cat_instructions = [inst for inst in instructions if inst['category_key'] == cat_key]
        
        ws_cat['A1'] = f"{cat_info['name']} 指令列表"
        ws_cat['A1'].font = Font(bold=True, size=14)
        ws_cat.merge_cells('A1:F1')
        
        for col, header in enumerate(headers[:6] + [headers[6]], 1):
            cell = ws_cat.cell(row=3, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, inst in enumerate(cat_instructions, 4):
            ws_cat.cell(row=row_idx, column=1, value=inst['name']).border = thin_border
            ws_cat.cell(row=row_idx, column=2, value=inst['funct6']).border = thin_border
            ws_cat.cell(row=row_idx, column=3, value=inst['funct3']).border = thin_border
            ws_cat.cell(row=row_idx, column=4, value=inst['instruction_type']).border = thin_border
            ws_cat.cell(row=row_idx, column=5, value=inst['operands']).border = thin_border
            ws_cat.cell(row=row_idx, column=6, value=inst['description_cn']).border = thin_border
            ws_cat.cell(row=row_idx, column=7, value=inst['description_en']).border = thin_border
        
        ws_cat.column_dimensions['A'].width = 15
        ws_cat.column_dimensions['B'].width = 20
        ws_cat.column_dimensions['C'].width = 8
        ws_cat.column_dimensions['D'].width = 25
        ws_cat.column_dimensions['E'].width = 20
        ws_cat.column_dimensions['F'].width = 35
        ws_cat.column_dimensions['G'].width = 30
    
    wb.save(EXCEL_FILE)
    print(f"Excel file generated: {EXCEL_FILE}")

def create_excel_basic():
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    instructions = data['instructions']
    
    with open(EXCEL_FILE.replace('.xlsx', '_basic.csv'), 'w', encoding='utf-8-sig') as f:
        f.write("指令名,分类,funct6,funct3,指令类型,操作数,中文描述,英文描述\n")
        for inst in instructions:
            f.write(f"{inst['name']},{inst['category']},{inst['funct6']},{inst['funct3']},")
            f.write(f"{inst['instruction_type']},{inst['operands']},{inst['description_cn']},{inst['description_en']}\n")
    
    print(f"Basic CSV file generated: {EXCEL_FILE.replace('.xlsx', '_basic.csv')}")

def main():
    print("Generating RISC-V V Vector Extension 0.7.1 Instruction Set Excel...")
    
    if PANDAS_AVAILABLE:
        create_excel_with_pandas()
    else:
        print("pandas/openpyxl not available, generating CSV instead...")
        create_excel_basic()

if __name__ == "__main__":
    main()
