#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RVV 1.0 Instruction Decode Status Update Script
Updates the JSON file with decode implementation status and generates Excel report.
"""

import json
import os
from datetime import datetime

# Define decode implementation status for each instruction
# Based on analysis of ct_idu_rf_pipe6_decd.v and ct_idu_rf_pipe7_decd.v

# Instructions implemented in both Pipe6 and Pipe7
BOTH_PIPE_IMPLEMENTED = {
    # Arithmetic
    'vadd', 'vsub', 'vrsub', 'vminu', 'vmin', 'vmaxu', 'vmax',
    # Logical
    'vand', 'vor', 'vxor',
    # Compare
    'vmseq', 'vmsne', 'vmsltu', 'vmslt', 'vmsleu', 'vmsle', 'vmsgtu', 'vmsgt',
    # Saturation
    'vsaddu', 'vsadd', 'vssubu', 'vssub', 'vaadd', 'vasub',
    # Shift
    'vsll', 'vsmul', 'vsrl', 'vsra', 'vssrl', 'vssra', 'vnsrl', 'vnsra', 'vnclipu', 'vnclip',
    # Move/Permutation
    'vrgather', 'vrgatherei16', 'vslideup', 'vslidedown', 'vslide1up', 'vslide1down', 'vcompress',
    # Carry/Borrow
    'vadc', 'vmadc', 'vsbc', 'vmsbc', 'vmerge', 'vmv',
    # Widening arithmetic
    'vwaddu', 'vwadd', 'vwsubu', 'vwsub', 'vwaddu.w', 'vwadd.w', 'vwsubu.w', 'vwsub.w',
    # Multiply
    'vmulhu', 'vmul', 'vmulhsu', 'vmulh', 'vwmulu', 'vwmulsu', 'vwmul',
    # Multiply-add
    'vmadd', 'vnmsub', 'vmacc', 'vnmsac', 'vwmaccu', 'vwmacc', 'vwmaccsu', 'vwmaccus',
    # Mask
    'vmandn', 'vmand', 'vmor', 'vmxor', 'vmorn', 'vmnand', 'vmnor', 'vmxnor',
    # FP arithmetic
    'vfadd', 'vfsub', 'vfrsub', 'vfmin', 'vfmax',
    # FP sign injection
    'vfsgnj', 'vfsgnjn', 'vfsgnjx',
    # FP multiply
    'vfmul', 'vfwmul',
    # FP multiply-add
    'vfmadd', 'vfnmadd', 'vfmsub', 'vfnmsub', 'vfmacc', 'vfnmacc', 'vfmsac', 'vfnmsac',
    'vfwmacc', 'vfwnmacc', 'vfwmsac', 'vfwnmsac',
    # FP compare
    'vmfeq', 'vmfle', 'vmflt', 'vmfne', 'vmfgt', 'vmfge',
    # FP class/merge
    'vfclass', 'vfmerge',
    # FP reduction
    'vfredusum', 'vfredosum', 'vfredmin', 'vfredmax',
    # FP widening
    'vfwadd', 'vfwsub', 'vfwadd.w', 'vfwsub.w',
    # Convert
    'vfcvt.xu.f.v', 'vfcvt.x.f.v', 'vfcvt.f.xu.v', 'vfcvt.f.x.v',
    'vfcvt.rtz.xu.f.v', 'vfcvt.rtz.x.f.v',
    'vfwcvt.xu.f.v', 'vfwcvt.x.f.v', 'vfwcvt.f.xu.v', 'vfwcvt.f.x.v', 'vfwcvt.f.f.v',
    'vfncvt.xu.f.w', 'vfncvt.x.f.w', 'vfncvt.f.xu.w', 'vfncvt.f.x.w', 'vfncvt.f.f.w',
    'vfncvt.rod.f.f.w', 'vfncvt.rtz.xu.f.w', 'vfncvt.rtz.x.f.w',
}

# Instructions implemented only in Pipe6
PIPE6_ONLY = {
    # Division
    'vdivu', 'vdiv', 'vremu', 'vrem',
    # Reduction
    'vredsum', 'vredand', 'vredor', 'vredxor', 'vredminu', 'vredmin', 'vredmaxu', 'vredmax',
    'vwredsumu', 'vwredsum',
    # FP division/sqrt
    'vfdiv', 'vfrdiv', 'vfsqrt',
    # FP special (RVV 1.0 new)
    'vfrsqrt7.v', 'vfrec7.v',
    # FP slide
    'vfslide1up', 'vfslide1down',
}

# Instructions not yet implemented or need verification
NOT_IMPLEMENTED = {
    # These are configuration or special instructions that may not need decode
    'vsetvli', 'vsetivli', 'vsetvl',
    # Atomic operations
    'vlxei', 'vsxei', 'vamoswapei', 'vamoaddei',
}

def get_decode_status(instruction_name, category_key):
    """Determine decode implementation status for an instruction."""
    # Normalize instruction name (remove variant suffixes)
    base_name = instruction_name.split('.')[0].replace('.v', '').replace('.vv', '').replace('.vx', '').replace('.vf', '').replace('.w', '').replace('.vi', '')
    
    # Check special cases
    if base_name in NOT_IMPLEMENTED or instruction_name in NOT_IMPLEMENTED:
        return 'not_implemented', 'N/A', 'N/A'
    
    # Check Pipe6 only
    if base_name in PIPE6_ONLY or instruction_name in PIPE6_ONLY:
        return 'implemented', 'Pipe6', 'FDSU/VREDU/VDIRU/FSPU'
    
    # Check both pipes
    if base_name in BOTH_PIPE_IMPLEMENTED or instruction_name in BOTH_PIPE_IMPLEMENTED:
        return 'implemented', 'Pipe6/Pipe7', 'VALU/VMULU/VSHIFT/VPERM/VMISC/FADD/FMAU/FSPU/FCNVT'
    
    # Default: need verification
    return 'need_verification', 'TBD', 'TBD'

def get_execution_unit(instruction_name, category_key):
    """Determine execution unit for an instruction."""
    eu_map = {
        'arithmetic': 'VALU',
        'logical': 'VALU',
        'compare': 'VALU',
        'shift': 'VSHIFT',
        'move': 'VPERM',
        'reduction': 'VREDU',
        'multiply-add': 'VMULU',
        'mask': 'VMISC',
        'fp': 'FADD/FMAU/FSPU/FDSU',
        'convert': 'FCNVT',
        'special': 'VMISC',
    }
    
    # Special cases
    if instruction_name in ['vdivu', 'vdiv', 'vremu', 'vrem']:
        return 'VDIRU'
    if instruction_name in ['vfdiv', 'vfrdiv', 'vfsqrt']:
        return 'FDSU'
    if instruction_name in ['vfrsqrt7.v', 'vfrec7.v']:
        return 'FSPU'
    if instruction_name in ['vredsum', 'vredand', 'vredor', 'vredxor', 'vredminu', 'vredmin', 'vredmaxu', 'vredmax']:
        return 'VREDU'
    if instruction_name in ['vwredsumu', 'vwredsum']:
        return 'VREDU'
    
    return eu_map.get(category_key, 'TBD')

def update_json_file(json_path):
    """Update JSON file with decode implementation status."""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Update metadata
    data['metadata']['decode_analysis_date'] = datetime.now().strftime('%Y-%m-%d')
    data['metadata']['decode_analysis_source'] = 'ct_idu_rf_pipe6_decd.v, ct_idu_rf_pipe7_decd.v'
    
    # Add decode status to each instruction
    for inst in data['instructions']:
        name = inst['name']
        category_key = inst.get('category_key', '')
        
        status, pipe, eu = get_decode_status(name, category_key)
        exec_unit = get_execution_unit(name, category_key)
        
        inst['decode_status'] = {
            'implemented': status,
            'decode_pipe': pipe,
            'execution_unit': exec_unit if eu == 'TBD' else eu,
            'analysis_date': datetime.now().strftime('%Y-%m-%d')
        }
    
    # Write updated JSON
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    return data

def create_excel_report(data, excel_path):
    """Create Excel report with instruction decode status."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call(['pip', 'install', 'openpyxl'])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RVV 1.0 Decode Status"
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    implemented_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    not_implemented_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    need_verification_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['序号', '指令名称', '类别', '指令类型', '操作数', '中文描述', 
               '译码状态', '译码Pipe', '执行单元', 'funct6', 'funct3', '备注']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Data rows
    row_num = 2
    for idx, inst in enumerate(data['instructions'], 1):
        decode_status = inst.get('decode_status', {})
        
        ws.cell(row=row_num, column=1, value=idx).border = thin_border
        ws.cell(row=row_num, column=2, value=inst['name']).border = thin_border
        ws.cell(row=row_num, column=3, value=inst.get('category', '')).border = thin_border
        ws.cell(row=row_num, column=4, value=inst.get('instruction_type', '')).border = thin_border
        ws.cell(row=row_num, column=5, value=inst.get('operands', '')).border = thin_border
        ws.cell(row=row_num, column=6, value=inst.get('description_cn', '')).border = thin_border
        
        # Decode status with color
        status_cell = ws.cell(row=row_num, column=7, value=decode_status.get('implemented', 'unknown'))
        status_cell.border = thin_border
        if decode_status.get('implemented') == 'implemented':
            status_cell.fill = implemented_fill
        elif decode_status.get('implemented') == 'not_implemented':
            status_cell.fill = not_implemented_fill
        else:
            status_cell.fill = need_verification_fill
        
        ws.cell(row=row_num, column=8, value=decode_status.get('decode_pipe', '')).border = thin_border
        ws.cell(row=row_num, column=9, value=decode_status.get('execution_unit', '')).border = thin_border
        ws.cell(row=row_num, column=10, value=inst.get('funct6', '')).border = thin_border
        ws.cell(row=row_num, column=11, value=inst.get('funct3', '')).border = thin_border
        
        # Notes
        diff_info = inst.get('diff_info', {})
        note = diff_info.get('details', '') if diff_info else ''
        ws.cell(row=row_num, column=12, value=note).border = thin_border
        
        row_num += 1
    
    # Adjust column widths
    column_widths = [6, 20, 25, 25, 25, 30, 12, 12, 20, 12, 10, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.cell(row=1, column=1, value="RVV 1.0 指令译码状态统计").font = Font(bold=True, size=14)
    
    # Count statistics
    total = len(data['instructions'])
    implemented = sum(1 for i in data['instructions'] if i.get('decode_status', {}).get('implemented') == 'implemented')
    not_impl = sum(1 for i in data['instructions'] if i.get('decode_status', {}).get('implemented') == 'not_implemented')
    need_verify = total - implemented - not_impl
    
    ws_summary.cell(row=3, column=1, value="总指令数:")
    ws_summary.cell(row=3, column=2, value=total)
    ws_summary.cell(row=4, column=1, value="已实现:")
    ws_summary.cell(row=4, column=2, value=implemented)
    ws_summary.cell(row=5, column=1, value="未实现:")
    ws_summary.cell(row=5, column=2, value=not_impl)
    ws_summary.cell(row=6, column=1, value="待验证:")
    ws_summary.cell(row=6, column=2, value=need_verify)
    ws_summary.cell(row=7, column=1, value="实现率:")
    ws_summary.cell(row=7, column=2, value=f"{implemented/total*100:.1f}%")
    
    # Analysis date
    ws_summary.cell(row=9, column=1, value="分析日期:")
    ws_summary.cell(row=9, column=2, value=datetime.now().strftime('%Y-%m-%d'))
    
    wb.save(excel_path)
    print(f"Excel report saved to: {excel_path}")

def main():
    """Main function."""
    base_path = r'd:\code\openc910\doc\Instruction_Set\riscv-v-spec-1.0'
    json_path = os.path.join(base_path, 'rvv_instructions.json')
    excel_path = os.path.join(base_path, 'rvv_instructions.xlsx')
    
    print("Updating JSON file with decode status...")
    data = update_json_file(json_path)
    print(f"JSON file updated: {json_path}")
    
    print("\nCreating Excel report...")
    create_excel_report(data, excel_path)
    
    print("\nDone!")

if __name__ == '__main__':
    main()
